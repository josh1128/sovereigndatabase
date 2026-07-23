"""Build chart_data.xlsx: one sheet per dashboard chart, mirroring app.py's parsing."""
import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import (AreaChart, BarChart, LineChart, PieChart,
                            Reference, Series)
from openpyxl.chart.marker import Marker
from openpyxl.chart.data_source import NumDataSource, NumRef
from openpyxl.drawing.colors import ColorChoice
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

SRC = "/mnt/user-data/uploads/data.xlsx"
OUT = "/home/claude/chart_data.xlsx"

# ── Parse exactly as app.py does ─────────────────────────────────────────────
raw = pd.read_excel(SRC, header=None)
DATA_START = 3
years, n_cols = [], 0
for c in range(DATA_START, raw.shape[1]):
    y = raw.iloc[0, c]
    if pd.isna(y):
        break
    s = str(y).strip()
    if s.endswith('p'):
        s = s[:-1]
    try:
        years.append(int(float(s)))
        n_cols += 1
    except (ValueError, TypeError):
        break


def row(idx):
    vals = raw.iloc[idx, DATA_START:DATA_START + n_cols].values
    out = []
    for v in vals:
        try:
            out.append(float(v) if pd.notna(v) and str(v).strip() != '****' else np.nan)
        except (ValueError, TypeError):
            out.append(np.nan)
    return pd.Series(out, index=years)


CREDITORS = {'IMF': 7, 'IBRD': 8, 'IDA': 9, 'IADB': 10, 'Paris Club': 11, 'China': 12,
             'Other official creditors': 13, 'FC bank loans': 14, 'FC bonds': 15,
             'Other private creditors': 16, 'LC debt': 17}
DEBTORS = {'Advanced economies': 23, 'Emerging-market and frontier economies': 24,
           'Heavily indebted poor countries': 25, 'Other developing economies': 26}
COUNTS = {'IMF': 41, 'IBRD': 42, 'IDA': 43, 'IADB': 44, 'Paris Club': 45, 'China': 46,
          'Other official creditors': 47, 'FC bank loans': 48, 'FC bonds': 49,
          'Other private creditors': 50, 'LC debt': 51}

total_debt = row(6)
total_sov = row(38)
LAST = max(y for y in years if pd.notna(total_debt[y]))

# Country block
countries = []
for i in range(66, len(raw)):
    iv = raw.iloc[i, 0]
    if pd.notna(iv) and isinstance(iv, (int, float)):
        name = str(raw.iloc[i, 1]).strip()
        vals = raw.iloc[i, DATA_START:DATA_START + n_cols].values
        ser = []
        for v in vals:
            try:
                ser.append(float(v) if pd.notna(v) and str(v).strip() != '****' else np.nan)
            except (ValueError, TypeError):
                ser.append(np.nan)
        countries.append((name, ser))
df_countries = pd.DataFrame([c[1] for c in countries],
                            index=[c[0] for c in countries], columns=years)

# ── Styling ──────────────────────────────────────────────────────────────────
FONT = "Arial"
H_FILL = PatternFill("solid", fgColor="1F3864")
H_FONT = Font(name=FONT, bold=True, color="FFFFFF", size=10)
TITLE_FONT = Font(name=FONT, bold=True, size=13, color="1F3864")
SUB_FONT = Font(name=FONT, italic=True, size=9, color="595959")
BODY = Font(name=FONT, size=10)
FORMULA_FONT = Font(name=FONT, size=10, color="008000")  # green = derived
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(bottom=THIN)

wb = Workbook()
wb.remove(wb.active)

# ── Chart palette (hex, matching the dashboard/published charts) ─────────────
CRED_HEX = {'IMF': '7B68A6', 'IBRD': 'FFFF00', 'IDA': 'E8112D', 'IADB': '3D7A99',
            'Paris Club': '1A7A3C', 'China': '2E75B6', 'Other official creditors': 'FFC000',
            'FC bank loans': 'FF9EDB', 'FC bonds': '8B6F47',
            'Other private creditors': '4CAF50', 'LC debt': '2B7A9E'}
DEBT_HEX = {'Advanced economies': 'E8112D',
            'Emerging-market and frontier economies': '4DC3E6',
            'Heavily indebted poor countries': '7030A0',
            'Other developing economies': 'FFC000'}


def style_chart(ch, title, y_title, x_title="Year", width=30, height=13):
    ch.title = title
    ch.y_axis.title = y_title
    ch.x_axis.title = x_title
    ch.width = width
    ch.height = height
    ch.style = 2
    # Keep axes visible (openpyxl can otherwise drop them in some viewers)
    ch.x_axis.delete = False
    ch.y_axis.delete = False
    return ch


def color_series(ch, hex_map, names):
    """Apply solid fills (and matching line colors) to each series in order."""
    for s, nm in zip(ch.series, names):
        hx = hex_map.get(nm)
        if not hx:
            continue
        s.graphicalProperties.solidFill = hx
        s.graphicalProperties.line.solidFill = hx


def add_cat_chart(ws, cls, first_col, last_col, n_rows, header_row, anchor,
                  title, y_title, grouping=None, overlap=None, hex_map=None,
                  names=None, marker=False, width=30, height=13):
    """Build a chart from a table laid out as Year in col A, series across."""
    ch = cls()
    data = Reference(ws, min_col=first_col, max_col=last_col,
                     min_row=header_row, max_row=header_row + n_rows)
    cats = Reference(ws, min_col=1, min_row=header_row + 1, max_row=header_row + n_rows)
    ch.add_data(data, titles_from_data=True)
    ch.set_categories(cats)
    if grouping:
        ch.grouping = grouping
        if overlap is not None:
            ch.overlap = overlap
    style_chart(ch, title, y_title, width=width, height=height)
    if hex_map and names:
        color_series(ch, hex_map, names)
    if marker:
        for s in ch.series:
            s.marker = Marker(symbol='none')
            s.smooth = False
    ws.add_chart(ch, anchor)
    return ch


def new_sheet(name, title, subtitle, note):
    ws = wb.create_sheet(name)
    ws["A1"] = title
    ws["A1"].font = TITLE_FONT
    ws["A2"] = subtitle
    ws["A2"].font = SUB_FONT
    ws["A3"] = f"Note: {note}"
    ws["A3"].font = SUB_FONT
    ws["A3"].alignment = Alignment(wrap_text=False)
    ws.freeze_panes = "B6"
    return ws


def write_table(ws, header_row, headers, index_vals, data_cols, numfmt="#,##0.0"):
    """Write a table with `Year` (or label) in col A starting at header_row."""
    for j, h in enumerate(headers):
        c = ws.cell(row=header_row, column=1 + j, value=h)
        c.font = H_FONT
        c.fill = H_FILL
        c.alignment = Alignment(horizontal="center", wrap_text=True, vertical="center")
    for i, idx in enumerate(index_vals):
        r = header_row + 1 + i
        c = ws.cell(row=r, column=1, value=idx)
        c.font = BODY
        c.border = BORDER
        if isinstance(idx, int):
            c.number_format = "0"
        for j, col in enumerate(data_cols):
            v = col[i]
            cc = ws.cell(row=r, column=2 + j,
                         value=(None if (isinstance(v, float) and np.isnan(v)) else v))
            cc.font = BODY
            cc.number_format = numfmt
            cc.border = BORDER
    ws.column_dimensions["A"].width = 30
    for j in range(len(headers) - 1):
        ws.column_dimensions[get_column_letter(2 + j)].width = 15
    ws.row_dimensions[header_row].height = 42


SRC_NOTE = "Source: BoC–BoE Sovereign Default Database (data.xlsx, July 22 2026 update)."

# ── README ───────────────────────────────────────────────────────────────────
ws = wb.create_sheet("README")
ws["A1"] = "Dashboard chart data — source and derivation"
ws["A1"].font = Font(name=FONT, bold=True, size=15, color="1F3864")
readme = [
    ("", ""),
    ("Purpose", "One sheet per chart in the Streamlit dashboard, holding exactly the series each chart plots."),
    ("Source file", "data.xlsx — 'Database of Sovereign Defaults', last update July 22 2026."),
    ("Parsing", "Data begins at spreadsheet column D (index 3); years are read from row 1 until the first"),
    ("", "blank column. A second year block follows that blank column and is deliberately ignored."),
    ("Projections", f"The final year ({LAST}) is flagged '2025p' in the source and is a projection."),
    ("Units", "Debt values are US$ millions in the source; charts divide by 1,000 to show US$ billions."),
    ("Green cells", "Values shown in green are live formulas (derived series), not pasted numbers."),
    ("Missing data", "Blank cells mean no data ('****' or empty in the source)."),
    ("", ""),
    ("Sheet", "Chart and source rows in data.xlsx (1-based row numbers)"),
    ("Chart 1", "Share of debt in default by creditor (pie) — rows 8–18, share of row 7 total."),
    ("Chart 2", "Default rates on FC bonds / bonds+bank loans — counts rows 49–50, sovereigns row 39."),
    ("Chart 3", "Total debt in default by creditor (stacked bars) — rows 8–18."),
    ("Chart 4", "Debt in default by debtor group (stacked bars) — rows 24–27."),
    ("Chart 5", "Proportion of debt in default by creditor (100% area) — rows 8–18, rescaled to 100%."),
    ("Chart 6", "Paris Club and China official loans in default — rows 12–13."),
    ("Chart 7", "Shares of global public debt and GDP — rows 35, 36, 37."),
    ("Chart 8", "Number of sovereign defaults by instrument — rows 49, 50, 52."),
    ("Map", "Debt in default by country — country block from row 67 onward."),
    ("", ""),
    ("Caveat — Chart 2", "The combined series sums bond and bank-loan defaulters; a sovereign in default"),
    ("", "on both instruments is therefore counted twice. The source has no union count."),
    ("Caveat — Chart 2", "Panel (a) of the published chart covers 1820–2020 using Suter (1992); that historical"),
    ("", "series is not in this workbook, so the dashboard shows the database's own period."),
]
for i, (k, v) in enumerate(readme, start=3):
    a = ws.cell(row=i, column=1, value=k)
    a.font = Font(name=FONT, bold=True, size=10)
    b = ws.cell(row=i, column=2, value=v)
    b.font = Font(name=FONT, size=10)
ws.column_dimensions["A"].width = 20
ws.column_dimensions["B"].width = 105

# ── Chart 1: pie shares (formulas) ───────────────────────────────────────────
ws = new_sheet("Chart 1", f"Chart 1: Total share of debt in default by creditor, {LAST}",
               "US$ millions and share of total.", SRC_NOTE)
names = list(CREDITORS)
ws.cell(row=6, column=1, value="Creditor").font = H_FONT
ws.cell(row=6, column=1).fill = H_FILL
for j, h in enumerate(["Debt in default (US$ mil)", "Share of total (%)"]):
    c = ws.cell(row=6, column=2 + j, value=h)
    c.font = H_FONT
    c.fill = H_FILL
    c.alignment = Alignment(horizontal="center", wrap_text=True, vertical="center")
for i, nm in enumerate(names):
    r = 7 + i
    ws.cell(row=r, column=1, value=nm).font = BODY
    v = row(CREDITORS[nm])[LAST]
    ws.cell(row=r, column=2, value=None if pd.isna(v) else float(v)).font = BODY
    ws.cell(row=r, column=2).number_format = "#,##0.0"
    f = ws.cell(row=r, column=3, value=f"=IFERROR(B{r}/$B${7+len(names)},0)")
    f.font = FORMULA_FONT
    f.number_format = "0.0%"
tr = 7 + len(names)
ws.cell(row=tr, column=1, value="Total").font = Font(name=FONT, bold=True, size=10)
ws.cell(row=tr, column=2, value=f"=SUM(B7:B{tr-1})").font = Font(name=FONT, bold=True, size=10, color="008000")
ws.cell(row=tr, column=2).number_format = "#,##0.0"
ws.cell(row=tr, column=3, value=f"=SUM(C7:C{tr-1})").font = Font(name=FONT, bold=True, size=10, color="008000")
ws.cell(row=tr, column=3).number_format = "0.0%"
ws.column_dimensions["A"].width = 32
ws.column_dimensions["B"].width = 24
ws.column_dimensions["C"].width = 18
ws.row_dimensions[6].height = 30

pie = PieChart()
pie.add_data(Reference(ws, min_col=2, min_row=6, max_row=6 + len(names)), titles_from_data=True)
pie.set_categories(Reference(ws, min_col=1, min_row=7, max_row=6 + len(names)))
pie.title = f"Chart 1: Share of debt in default by creditor, {LAST}"
pie.width, pie.height = 22, 14
pie.dataLabels = None
ws.add_chart(pie, "E6")

# ── Chart 2: default rates (formulas from counts) ────────────────────────────
ws = new_sheet("Chart 2", "Chart 2: Sovereign default rates on FC bonds and bank loans",
               "Counts of sovereigns in default, and rates as a share of all sovereigns.",
               "Combined series sums bond and bank-loan defaulters (double-counts sovereigns in default on both). " + SRC_NOTE)
hdr = ["Year", "Sovereigns in default: FC bonds", "Sovereigns in default: FC bank loans",
       "Total sovereigns", "FC bonds (%)", "FC bonds and bank loans (%)"]
for j, h in enumerate(hdr):
    c = ws.cell(row=6, column=1 + j, value=h)
    c.font = H_FONT
    c.fill = H_FILL
    c.alignment = Alignment(horizontal="center", wrap_text=True, vertical="center")
b_ct, l_ct = row(COUNTS['FC bonds']), row(COUNTS['FC bank loans'])
for i, y in enumerate(years):
    r = 7 + i
    ws.cell(row=r, column=1, value=y).font = BODY
    ws.cell(row=r, column=1).number_format = "0"
    for j, ser in enumerate([b_ct, l_ct, total_sov]):
        v = ser[y]
        cc = ws.cell(row=r, column=2 + j, value=None if pd.isna(v) else float(v))
        cc.font = BODY
        cc.number_format = "#,##0"
    f1 = ws.cell(row=r, column=5, value=f"=IFERROR(B{r}/D{r},\"\")")
    f2 = ws.cell(row=r, column=6, value=f"=IFERROR((B{r}+C{r})/D{r},\"\")")
    for f in (f1, f2):
        f.font = FORMULA_FONT
        f.number_format = "0.0%"
ws.column_dimensions["A"].width = 10
for col in "BCDEF":
    ws.column_dimensions[col].width = 20
ws.row_dimensions[6].height = 44

add_cat_chart(ws, LineChart, 5, 6, len(years), 6, "H6",
              "Chart 2: Sovereign default rates on FC bonds and bank loans",
              "% of all sovereigns", marker=True)

# ── Chart 3: debt by creditor ────────────────────────────────────────────────
ws = new_sheet("Chart 3", "Chart 3: Total sovereign debt in default by creditor",
               "US$ millions (dashboard plots US$ billions = these values / 1,000).",
               "LC is local currency, FC is foreign currency. " + SRC_NOTE)
cols = [[row(v)[y] for y in years] for v in CREDITORS.values()]
write_table(ws, 6, ["Year"] + names, years, cols)
add_cat_chart(ws, BarChart, 2, 1 + len(names), len(years), 6, "N6",
              "Chart 3: Total sovereign debt in default by creditor",
              "US$ millions", grouping="stacked", overlap=100,
              hex_map=CRED_HEX, names=names, width=34, height=15)

# ── Chart 4: debt by debtor ──────────────────────────────────────────────────
ws = new_sheet("Chart 4", "Chart 4: Sovereign debt in default by debtor",
               "US$ millions (dashboard plots US$ billions).", SRC_NOTE)
dcols = [[row(v)[y] for y in years] for v in DEBTORS.values()]
write_table(ws, 6, ["Year"] + list(DEBTORS), years, dcols)
add_cat_chart(ws, BarChart, 2, 1 + len(DEBTORS), len(years), 6, "G6",
              "Chart 4: Sovereign debt in default by debtor",
              "US$ millions", grouping="stacked", overlap=100,
              hex_map=DEBT_HEX, names=list(DEBTORS), width=34, height=15)

# ── Chart 5: proportions (formulas) ──────────────────────────────────────────
ws = new_sheet("Chart 5", "Chart 5: Proportion of debt in default by creditor",
               "Each creditor as a share of total debt in default (%).",
               "Shares are computed from the Chart 3 levels and sum to 100% each year. " + SRC_NOTE)
for j, h in enumerate(["Year"] + names):
    c = ws.cell(row=6, column=1 + j, value=h)
    c.font = H_FONT
    c.fill = H_FILL
    c.alignment = Alignment(horizontal="center", wrap_text=True, vertical="center")
ncred = len(names)
for i, y in enumerate(years):
    r = 7 + i
    ws.cell(row=r, column=1, value=y).font = BODY
    ws.cell(row=r, column=1).number_format = "0"
    for j in range(ncred):
        src_col = get_column_letter(2 + j)
        f = ws.cell(row=r, column=2 + j,
                    value=f"=IFERROR('Chart 3'!{src_col}{r}/SUM('Chart 3'!$B{r}:$L{r}),\"\")")
        f.font = FORMULA_FONT
        f.number_format = "0.0%"
ws.column_dimensions["A"].width = 10
for j in range(ncred):
    ws.column_dimensions[get_column_letter(2 + j)].width = 15
ws.row_dimensions[6].height = 42

add_cat_chart(ws, AreaChart, 2, 1 + ncred, len(years), 6, "N6",
              "Chart 5: Proportion of debt in default by creditor",
              "% of total debt in default", grouping="percentStacked",
              hex_map=CRED_HEX, names=names, width=34, height=15)

# ── Chart 6: Paris Club & China ──────────────────────────────────────────────
ws = new_sheet("Chart 6", "Chart 6: Official loans in default for Paris Club and China",
               "US$ millions (dashboard plots US$ billions).", SRC_NOTE)
write_table(ws, 6, ["Year", "Paris Club", "China"], years,
            [[row(11)[y] for y in years], [row(12)[y] for y in years]])
add_cat_chart(ws, BarChart, 2, 3, len(years), 6, "F6",
              "Chart 6: Official loans in default for Paris Club and China",
              "US$ millions", grouping="stacked", overlap=100,
              hex_map={'Paris Club': 'FF0000', 'China': '5B9BD5'},
              names=['Paris Club', 'China'], width=32, height=14)

# ── Chart 7: shares of debt and GDP ──────────────────────────────────────────
ws = new_sheet("Chart 7", "Chart 7: Debt in default as a share of global public debt and GDP",
               "Percent. Nominal GDP is used.", SRC_NOTE)
write_table(ws, 6,
            ["Year", "Share of global public debt (%)", "Share of world GDP (%)",
             "Share of EM/other developing GDP (%)"],
            years,
            [[row(34)[y] for y in years], [row(36)[y] for y in years],
             [row(35)[y] for y in years]], numfmt="0.00")
add_cat_chart(ws, LineChart, 2, 4, len(years), 6, "G6",
              "Chart 7: Debt in default as a share of global public debt and GDP",
              "%", hex_map={'Share of global public debt (%)': 'A52929',
                            'Share of world GDP (%)': '2E9BD6',
                            'Share of EM/other developing GDP (%)': 'F5A623'},
              names=["Share of global public debt (%)", "Share of world GDP (%)",
                     "Share of EM/other developing GDP (%)"],
              marker=True, width=32, height=14)

# ── Chart 8: number of defaults ──────────────────────────────────────────────
ws = new_sheet("Chart 8", "Chart 8: Number of sovereign defaults by instrument",
               "Count of sovereigns in default.",
               "FC is foreign currency, LC is local currency. " + SRC_NOTE)
write_table(ws, 6, ["Year", "FC bank loans", "FC bonds", "LC debt"], years,
            [[row(48)[y] for y in years], [row(49)[y] for y in years],
             [row(51)[y] for y in years]], numfmt="#,##0")
add_cat_chart(ws, LineChart, 2, 4, len(years), 6, "G6",
              "Chart 8: Number of sovereign defaults by instrument",
              "Number of sovereigns",
              hex_map={'FC bank loans': '2E75B6', 'FC bonds': 'C0504D', 'LC debt': '9BBB59'},
              names=["FC bank loans", "FC bonds", "LC debt"],
              marker=True, width=32, height=14)

# ── Map: country data ────────────────────────────────────────────────────────
ws = new_sheet("Map", "Map: Total debt in default by country",
               "US$ millions, by country and year.",
               "Grey on the map = no data or zero. " + SRC_NOTE)
ccols = [[df_countries.loc[nm, y] for nm in df_countries.index] for y in years]
write_table(ws, 6, ["Country"] + [str(y) for y in years],
            list(df_countries.index), ccols)
ws.column_dimensions["A"].width = 36

# A 166-row chart is unreadable, so chart the top 15 defaulters in the latest year.
top = df_countries[LAST].dropna()
top = top[top > 0].sort_values(ascending=False).head(15)
start = 6 + len(df_countries) + 3
ws.cell(row=start - 1, column=1,
        value=f"Top 15 countries by debt in default, {LAST} (US$ millions)").font = TITLE_FONT
for j, h in enumerate(["Country", f"Debt in default {LAST}"]):
    c = ws.cell(row=start, column=1 + j, value=h)
    c.font = H_FONT
    c.fill = H_FILL
    c.alignment = Alignment(horizontal="center", wrap_text=True, vertical="center")
for i, (nm, v) in enumerate(top.items()):
    ws.cell(row=start + 1 + i, column=1, value=nm).font = BODY
    cc = ws.cell(row=start + 1 + i, column=2, value=float(v))
    cc.font = BODY
    cc.number_format = "#,##0.0"

bar = BarChart()
bar.type = "bar"
bar.add_data(Reference(ws, min_col=2, min_row=start, max_row=start + len(top)),
             titles_from_data=True)
bar.set_categories(Reference(ws, min_col=1, min_row=start + 1, max_row=start + len(top)))
style_chart(bar, f"Map: Top 15 countries by debt in default, {LAST}",
            "US$ millions", x_title="Country", width=26, height=15)
bar.legend = None
for s_ in bar.series:
    s_.graphicalProperties.solidFill = "E87040"
ws.add_chart(bar, f"D{start}")

wb.save(OUT)
print("saved", OUT)
print("sheets:", wb.sheetnames)
print("years:", years[0], "->", years[-1], "| countries:", len(df_countries))
